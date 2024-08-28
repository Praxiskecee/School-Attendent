import cv2
import sqlite3
import pathlib
import os
import tkinter as tk
from datetime import datetime, date
from PIL import Image, ImageTk
import numpy as np
import face_recognition
from openpyxl import Workbook, load_workbook
import pyttsx3

# Paths
current_dir = pathlib.Path(__file__).parent
img_dir = current_dir.joinpath('img')
screenshots_dir = current_dir.joinpath('screenshots')
morning_screenshots_dir = screenshots_dir.joinpath('morning')
afternoon_screenshots_dir = screenshots_dir.joinpath('afternoon')
db_path = current_dir.joinpath('faces.db')
spreadsheet_path = current_dir.joinpath('face_data.xlsx')

# Create directories if not exist
img_dir.mkdir(parents=True, exist_ok=True)
screenshots_dir.mkdir(parents=True, exist_ok=True)
morning_screenshots_dir.mkdir(parents=True, exist_ok=True)
afternoon_screenshots_dir.mkdir(parents=True, exist_ok=True)

# Create or load the spreadsheet
if spreadsheet_path.exists():
    workbook = load_workbook(spreadsheet_path)
else:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Face Data"
    sheet.append(["Name", "Role", "Image Path", "Log", "Time"])  # Header row
    workbook.save(spreadsheet_path)

# Connect to SQLite database
conn = sqlite3.connect(db_path)
cursor = conn.cursor()

# Create table for storing face data
cursor.execute('''
CREATE TABLE IF NOT EXISTS faces (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT,
    role TEXT,
    image_path TEXT,
    encoding BLOB,
    log TEXT
)
''')
conn.commit()

# Initialize Text-to-Speech engine
engine = pyttsx3.init()

def speak(text):
    engine.say(text)
    engine.runAndWait()

def save_to_spreadsheet(name, role, img_path, log):
    workbook = load_workbook(spreadsheet_path)
    sheet = workbook.active
    current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    sheet.append([name, role, img_path, log, current_time])
    workbook.save(spreadsheet_path)

def recognize_face(face_encoding):
    cursor.execute("SELECT * FROM faces")
    rows = cursor.fetchall()
    for row in rows:
        known_encoding = np.frombuffer(row[4], dtype=np.float64)
        matches = face_recognition.compare_faces([known_encoding], face_encoding, tolerance=0.4)
        if True in matches:
            return row
    return None

def save_identity(name, face, role, face_encoding):
    timestamp = datetime.now().timestamp()
    img_path = img_dir.joinpath(f'{int(timestamp)}.jpg')
    cv2.imwrite(str(img_path), face)

    encoding_blob = face_encoding.tobytes()
    cursor.execute('''
    INSERT INTO faces (name, role, image_path, encoding, log) VALUES (?, ?, ?, ?, ?)
    ''', (name, role, str(img_path), encoding_blob, '[]'))
    conn.commit()

    save_to_spreadsheet(name, role, str(img_path), '[]')
    speak(f"Wajah {name} telah terdaftar.")

def register_new_face(frame, face, face_encoding):
    global register_mode
    register_mode = True

    root = tk.Tk()
    root.title("Daftarkan Wajah")

    tk.Label(root, text='Masukkan nama:').pack()
    name_entry = tk.Entry(root)
    name_entry.pack()

    tk.Label(root, text='Masukkan role:').pack()
    role_entry = tk.Entry(root)
    role_entry.pack()

    face_img = Image.fromarray(cv2.cvtColor(face, cv2.COLOR_BGR2RGB))
    face_img = ImageTk.PhotoImage(image=face_img)
    tk.Label(root, image=face_img).pack()

    def save():
        name = name_entry.get()
        role = role_entry.get()
        save_identity(name, face, role, face_encoding)
        root.destroy()
        global register_mode
        register_mode = False  # Reset the registration mode after saving

    tk.Button(root, text='Simpan', command=save).pack()
    root.mainloop()

def log_attendance(entry_id):
    today = date.today().isoformat()
    now = datetime.now().isoformat()

    cursor.execute("SELECT log FROM faces WHERE id = ?", (entry_id,))
    log = eval(cursor.fetchone()[0])

    if len(log) == 0 or log[-1]['date'] != today:
        log.append({
            'date': today,
            'attendance_time': now,
            'go_home_time': None
        })
    else:
        log[-1]['go_home_time'] = now

    cursor.execute("UPDATE faces SET log = ? WHERE id = ?", (str(log), entry_id))
    conn.commit()

def get_screenshot_folder():
    now = datetime.now()
    current_time = now.time()
    morning_start = datetime.strptime('07:30:00', '%H:%M:%S').time()
    morning_end = datetime.strptime('12:00:00', '%H:%M:%S').time()
    afternoon_start = datetime.strptime('14:00:00', '%H:%M:%S').time()
    night_end = datetime.strptime('18:00:00', '%H:%M:%S').time()

    if morning_start <= current_time < morning_end:
        return morning_screenshots_dir
    elif afternoon_start <= current_time or current_time < night_end:
        return afternoon_screenshots_dir
    else:
        return None

# Function to determine the greeting based on the time of day
def get_greeting():
    now = datetime.now().time()
    morning_start = datetime.strptime('07:30:00', '%H:%M:%S').time()
    morning_end = datetime.strptime('12:00:00', '%H:%M:%S').time()
    afternoon_start = datetime.strptime('14:00:00', '%H:%M:%S').time()
    afternoon_end = datetime.strptime('18:00:00', '%H:%M:%S').time()

    if morning_start <= now < morning_end:
        return "Selamat datang"
    elif afternoon_start <= now < afternoon_end:
        return "Selamat jalan"
    else:
        return None  # No greeting outside specified hours

# Try different VideoCapture backends
backends = [cv2.CAP_MSMF, cv2.CAP_VFW, cv2.CAP_DSHOW]
cap = None

for backend in backends:
    cap = cv2.VideoCapture(0, backend)
    if cap.isOpened():
        print(f"Successfully opened camera with backend: {backend}")
        break
else:
    print("Failed to open camera with any backend.")
    exit()

screenshot_taken_for_face = {}  # Dictionary to track if screenshot is taken for a specific face
register_mode = False

while True:
    ret, frame = cap.read()
    if not ret:
        print("Failed to grab frame.")
        break

    frame = cv2.convertScaleAbs(frame, alpha=1.3, beta=30)

    rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    face_locations = face_recognition.face_locations(rgb_frame)
    face_encodings = face_recognition.face_encodings(rgb_frame, face_locations)

    for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
        face_img = frame[top:bottom, left:right]
        face_key = tuple(np.round(face_encoding, decimals=5))  # Round to avoid floating-point issues

        if register_mode:
            # Skip recognition during registration mode
            cv2.putText(frame, "Mode Registrasi Aktif", (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 255), 2)
            continue

        entry = recognize_face(face_encoding)
        if entry:
            log_attendance(entry[0])
            name_and_role = f"{entry[1]} ({entry[2]})"
            cv2.putText(frame, name_and_role, (left, top-10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (255, 255, 255), 2)
            
            if face_key not in screenshot_taken_for_face:  # Only take screenshot if not taken for this face
                screenshot_frame = frame.copy()
                detection_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                cv2.putText(screenshot_frame, detection_time, (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
                screenshot_folder = get_screenshot_folder()
                
                if screenshot_folder:
                    screenshot_path = screenshot_folder.joinpath(f'{entry[1]}_{datetime.now().strftime("%Y%m%d%H%M%S")}.jpg')
                    cv2.imwrite(str(screenshot_path), screenshot_frame)
                    screenshot_taken_for_face[face_key] = True  # Mark that screenshot has been taken for this face

            # Speak when a recognized face is detected, with a greeting based on time
            greeting = get_greeting()
            if greeting:
                speak(f"{greeting}, {entry[1]}!")
        else:
            cv2.putText(frame, "Tekan 's' untuk mendaftarkan wajah baru", (left, top-10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 0, 255), 2)

        cv2.rectangle(frame, (left, top), (right, bottom), (255, 0, 0), 2)

    # Display current time on the frame
    current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    cv2.putText(frame, current_time, (10, 60), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)

    cv2.putText(frame, 'Tekan [q] untuk keluar', (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)
    cv2.imshow('Face Recognition', frame)

    key = cv2.waitKey(1) & 0xFF
    if key == ord('q'):
        break
    elif key == ord('s'):
        if face_encodings:
            register_new_face(frame, face_img, face_encoding)
            screenshot_taken_for_face.clear()  # Clear screenshot tracking when registering new face

cap.release()
cv2.destroyAllWindows()
conn.close()
